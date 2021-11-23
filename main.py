# MIT License
#
# Copyright (c) 2021 GGggg-sp
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import numpy as np
import pickle as pkl
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill

# The number of nodes on the map
number_of_nodes = 92

# The tools collection list
tools_list = ['手', '藤条', '小红', '小绿', '树脂棍', '皮带', '戒尺']

# The positions and tools for spanking:
# The positions are used as the keys in the dictionary and you can specific if there are some tools not suitable for
# some positions, add your favorite position & tools as you like.
position_tools_dict = {'OTK': [i for i in tools_list if i not in ['皮带', '小绿', '藤条']],  # for example: pidai, xiaolv
                       # and tentiao are not suitable for otk spanking, so we remove them by the expression above
                       '趴在床沿': [i for i in tools_list if i not in ['手']],
                       '床上平趴垫枕头': [i for i in tools_list if i not in ['手']],
                       '身体趴在桌子上': [i for i in tools_list if i not in []],
                       '手扶墙壁': [i for i in tools_list if i not in []],
                       '手握脚踝': [i for i in tools_list if i not in []],
                       '跪撅': [i for i in tools_list if i not in []],
                       '换尿布式': [i for i in tools_list if i not in ['皮带']]}

# some novel counting patterns that the spankee must count the numbers using bases such as dec, oct, etc
# you can control the possibilities of each base as long as they enjoy a sum of 1
base_list = ['八进制', '二进制', '十进制', '十六进制']
possibilities_base_list = [0.15, 0.05, 0.7, 0.1]

# not sp event list contains all the event types that are not spanking or special game patterns in 404 sanctuary
not_sp_event_list = ['揉揉', '晾臀', '后退一格', '后退两格', '回到起点', '前进一格', '前进两格', '请主随意']
# special events featured by some dalaos in 404 sanctuary
special_event_list = ['地刺王冰冰出现了！请您后退六格！', 'Genecro 请您回到起点并请罚', 'JayZ 为您点了数量任意的藤条',
                      'Zara 团宠觉得您应该享受更多的小红', '礼物姐看了眼你的屁屁表示不屑并建议增加任意数量的巴掌',
                      '等子姐在等您前往避难所，因此您增加了晾臀时间',
                      '泰坦：您今日戒色了吗？请您根据主的要求 DIY ', '冷子姐请您喝奶啤哦！喝完增加任意数量的棍状工具',
                      '小诺姐值班时看到了您并将您写进了小说，本轮您可以自由选择将受到的惩罚',
                      'linu 告诉您反被为主并不是好的选择，并为您增加了亿些皮带数目', 'ladboy 的藤条到货了想请您去品尝，因此您增加了任意数量的藤条！',
                      '悠久のSP資料館欢迎您前来阅览，您获得了五分钟的休息时间！',
                      '火球太太声称火球在拍灰，火球对此事并没有回应，因此太太决定为您增加重度戒尺',
                      'vv 最近失踪辣，可能是在躲着姐姐叭，请您帮助 vv 承担亿些 OTK 手掌',
                      '你这个一点都不重度诶，拳王如实说道，并为您点了重度藤条',
                      'Netspanker对你主的手法表示质疑，您有一次机会可以反被为主，冲！',
                      'dokoham 永远在默默无闻的搬运着资源，您选择挨任意数量的板状工具支持他']
special_event_list_bac = special_event_list.copy()
# possibilities for sp, not sp event and the special event
possibilities_event_list = [0.75, 0.15, 0.1]
event_list = ['sp', 'not sp', 'special']

num_max_per_epoch = 40
num_list = list(range(10, num_max_per_epoch))


# The argument process_p is to indicate the process of the events.
# For example, we can control the events possibility distribution in different stages of the whole spanking ludo games.
def generate_single_event(process_p):
    event_type = np.random.choice(event_list, 1, p=possibilities_event_list).tolist()[0]
    if event_type == 'sp':
        position = np.random.choice(list(position_tools_dict.keys()), 1).tolist()[0]
        tool = np.random.choice(position_tools_dict[position], 1).tolist()[0]
        base = np.random.choice(base_list, 1, p=possibilities_base_list).tolist()[0]
        num = str(np.random.choice(num_list, 1).tolist()[0])

        event_str = '姿势：' + position + '\n工具：' + tool + '\n数目：' + str(num) + '\n以' + base + '报数'
        event_dict = {'type': 'spanking', 'position': position, 'tool': tool, 'base': base, 'num': num}
    elif event_type == 'not sp':
        event_str = np.random.choice(not_sp_event_list)
        event_dict = {'type': 'not spanking', 'content': event_str}
    else:
        event_str = np.random.choice(special_event_list)
        special_event_list.remove(event_str)
        event_dict = {'type': 'special', 'content': event_str}
    return event_str, event_dict

# Main procedure for generate events
def generate_event_series(event_max_num: int):
    events_list = []
    events_str_list = []
    for current_event_num in range(event_max_num):
        current_event_str, current_event_dict = generate_single_event(current_event_num / event_max_num)
        events_str_list.append(current_event_str)
    return events_str_list, events_list


def save_excel(event_series_str, excel_filename):
    workbook = Workbook()
    sheet = workbook.active

    # Calculate the size of the ludo map
    current_max_h = int(np.ceil(np.sqrt(len(event_series_str) * 2)) + 2)
    current_min_h = 1
    current_max_v = current_max_h
    current_min_v = 1
    current_h = 1
    current_v = 1
    speed_h = 1
    speed_v = 0
    base_ascii = 64
    space = 2

    for i in range(1, current_max_v):
        sheet.row_dimensions[i].height = 75
    for i in range(1, current_max_h):
        sheet.column_dimensions[chr(base_ascii+i)].width = 20
    for ind in range(len(event_series_str)):
        current_h = current_h + speed_h
        current_v = current_v + speed_v
        current_index = chr(base_ascii + current_v) + str(current_h)
        sheet[current_index] = event_series_str[ind]
        center_aligned_text = Alignment(horizontal="center", wrapText=True, vertical="center")
        double_border_side = Side(border_style="double")
        square_border = Border(top=double_border_side,
                               right=double_border_side,
                               bottom=double_border_side,
                               left=double_border_side)

        sheet[current_index].border = square_border
        sheet[current_index].alignment = center_aligned_text
        if event_series_str[ind] in special_event_list_bac:
            sheet[current_index].fill = PatternFill(start_color="F48225", fill_type='solid')
        elif event_series_str[ind] in not_sp_event_list:
            sheet[current_index].fill = PatternFill(start_color="BBBBAA", fill_type='solid')
        # check if reached the boundary
        # ->
        if current_h + space == current_max_h and speed_h > 0:
            speed_h = 0
            speed_v = 1
            current_max_h = current_max_h - space
        # |
        # v
        if current_v + space == current_max_v and speed_v > 0:
            speed_h = -1
            speed_v = 0
            current_max_v = current_max_v - space
        # <-
        if current_h - space == current_min_h and speed_h < 0:
            speed_h = 0
            speed_v = -1
            current_min_h = current_min_h + space
        # ^
        # |
        if current_v - space == current_min_v and speed_v < 0:
            speed_h = 1
            speed_v = 0
            current_min_v = current_min_v + space

        workbook.save(excel_filename)
    return


if __name__ == '__main__':

    event_series_str, event_series_list = generate_event_series(event_max_num=number_of_nodes)
    # for i in range(len(event_series_str)):
    #     print('第', i+1, '格')
    #     print(event_series_str[i])

    # Save generated event list to disk for further usage
    with open('sp_ludo.pkl', 'wb') as f:
        pkl.dump(event_series_list, f)

    # Save the excel file
    save_excel(event_series_str, "sp ludo.xlsx")
    print('Excel file saved!')
